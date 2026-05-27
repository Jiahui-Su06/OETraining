import unicodedata
from PySide6.QtGui import QPainter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ui_mainwindow import Ui_MainWindow
from core import ViewGraphicsScene, LaserResonator
from PySide6.QtCore import Slot
from PySide6.QtWidgets import (
    QMainWindow, QMessageBox,
    QFileDialog,
)


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        # set default value
        self.wavelength_DSpinBox.setValue(1.046)
        self.cavity_length_DSpinBox.setValue(450.0)
        self.left_radius_DSpinBox.setValue(500.0)
        self.right_radius_DSpinBox.setValue(500.0)
        self.key_step_DSpinBox.setValue(20.0)
        self.d1Box.setValue(20.0)
        self.d2Box.setValue(100.0)
        self.fBox.setValue(40.0)
        self.comboBox.addItems(
            ["None", "F810FC-635", "F810FC-850", "F810FC-1064"]
        )
        self.comboBox.setCurrentIndex(0)
        
        self.scene = ViewGraphicsScene(self)
        self.graphicsView.setScene(self.scene)
        self.graphicsView.setRenderHint(QPainter.Antialiasing)

        wl = self.wavelength_DSpinBox.value()
        L = self.cavity_length_DSpinBox.value()
        r1 = self.left_radius_DSpinBox.value()
        r2 = self.right_radius_DSpinBox.value()
        d1 = self.d1Box.value()
        d2 = self.d2Box.value()
        f = self.fBox.value()
        lens = self.checkBox.isChecked()

        self.resonator = LaserResonator(
            parent_scene=self.scene,
            wl=wl, L=L, r1=r1, r2=r2,
            d1=d1, d2=d2, f=f, lens=lens
        )

        self.wavelength_DSpinBox.valueChanged.connect(self.redraw_scene)
        self.cavity_length_DSpinBox.valueChanged.connect(self.redraw_scene)
        self.left_radius_DSpinBox.valueChanged.connect(self.redraw_scene)
        self.right_radius_DSpinBox.valueChanged.connect(self.redraw_scene)
        self.checkBox.toggled.connect(self.redraw_scene)
        self.d1Box.valueChanged.connect(self.redraw_scene)
        self.d2Box.valueChanged.connect(self.redraw_scene)
        self.fBox.valueChanged.connect(self.redraw_scene)
        self.comboBox.currentTextChanged.connect(self.redraw_scene)
        self.help_Button.clicked.connect(self.show_help_dialog)
        self.reset_Button.clicked.connect(self.reset_Button_clicked)

        self.actionHelp.triggered.connect(self.show_help_dialog)
        self.actionQuit.triggered.connect(self.close)
        self.actionSave.triggered.connect(self.saveData)
        self.actionHelp.triggered.connect(self.show_help_dialog)

        self.textBrowser.append(
            "> <span style='color: green'>Initialized successfully</span>\n"
        )


    def zoom_in(self):
        self.graphicsView.scale(2, 2)


    def zoom_out(self):
        self.graphicsView.scale(0.5, 0.5)


    @Slot()
    def redraw_scene(self):
        wl = self.wavelength_DSpinBox.value()
        L = self.cavity_length_DSpinBox.value()
        r1 = self.left_radius_DSpinBox.value()
        r2 = self.right_radius_DSpinBox.value()
        d1 = self.d1Box.value()
        d2 = self.d2Box.value()
        lens = self.checkBox.isChecked()
        if self.comboBox.currentText() == "None":
            self.fBox.setReadOnly(False)
            f = self.fBox.value()
        elif self.comboBox.currentText() == "F810FC-635":
            self.fBox.setReadOnly(True)
            f = 35.41  # mm
            self.fBox.setValue(35.41)
        elif self.comboBox.currentText() == "F810FC-850":
            self.fBox.setReadOnly(True)
            f = 36.60  # mm
            self.fBox.setValue(36.60)
        elif self.comboBox.currentText() == "F810FC-1064":
            self.fBox.setReadOnly(True)
            f = 36.20  # mm
            self.fBox.setValue(36.20)

        self.resonator.update_params(
            wl=wl, L=L, r1=r1, r2=r2,
            d1=d1, d2=d2, f=f, lens=lens
        )
        z_waist_from_lens, waist_width = self.resonator.get_external_waist()
        self.wasitEdit.setText(f"{waist_width:.2f}")
        self.distanceEdit.setText(f"{z_waist_from_lens:.2f}")
        self.resonator.redraw()


    @Slot()
    def show_help_dialog(self):
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Function Key Introduction")
        msg_box.setText(
            "1. Use the left/right arrow keys to adjust cavity length.\n" \
            "2. Use the up/down arrow keys to adjust the radius of the right mirror.\n" \
            "3. Press the W/S keys to adjust the radius of the left mirror.\n" \
            "4. Hold Ctrl and scroll the mouse wheel to zoom the view.\n" \
            "5. Drag the left mirror to move the entire resonator.")
        msg_box.setIcon(QMessageBox.Information)

        msg_box.exec()


    @Slot()
    def reset_Button_clicked(self):
        self.wavelength_DSpinBox.setValue(1.064)
        self.cavity_length_DSpinBox.setValue(450)
        self.left_radius_DSpinBox.setValue(500)
        self.right_radius_DSpinBox.setValue(500)
        self.key_step_DSpinBox.setValue(20)

        self.redraw_scene()
    

    def saveData(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Resonator Parameters",
            "docs/",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return
        
        params = [
            ("Wavelength (μm)", self.wavelength_DSpinBox.value()),
            ("Cavity Length (mm)", self.cavity_length_DSpinBox.value()),
            ("Left Mirror Radius (mm)", self.left_radius_DSpinBox.value()),
            ("Right Mirror Radius (mm)", self.right_radius_DSpinBox.value()),
            ("Fitted with a collimating lens", "Yes" if self.checkBox.isChecked() else "No"),
            ("Lens Part Number", self.comboBox.currentText()),
            ("Focal Length (mm)", self.fBox.value()),
            ("Gap Length (mm)", self.d1Box.value()),
            ("View Length (mm)", self.d2Box.value()),
            ("Waist Width (mm)", round(float(self.wasitEdit.text().strip()), 2)),
            ("Waist From Lens (mm)",  round(float(self.distanceEdit.text().strip()), 2)),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Resonator Parameters"

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='0070C0')
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_gray = Side(style='thin', color='D9D9D9')
        h_border = Border(bottom=thin_gray)

        headers = ["Parameter Name", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = h_border

        for row_idx, (param_name, param_value) in enumerate(params, 2):
            name_cell = ws.cell(row=row_idx, column=1, value=param_name)
            name_cell.alignment = left_align
            name_cell.border = h_border

            value_cell = ws.cell(row=row_idx, column=2, value=param_value)
            value_cell.alignment = right_align
            value_cell.border = h_border

        def display_width(text):
            return sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(text or ''))
        
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_width = max((display_width(c.value) for c in col_cells if c.value is not None), default=0)
            ws.column_dimensions[letter].width = max(8, min(max_width * 1.1 + 3, 50))

        try:
            wb.save(file_path)
            self.textBrowser.append(
                f"> <span style='color: green'>Parameters saved to:\n{file_path}</span>\n"
            )
        except Exception as e:
            self.textBrowser.append(
                f"> <span style='color: red'>Save failed:\n{str(e)}</span>\n"
            )
            QMessageBox.critical(self, "Error", f"Save failed:\n{str(e)}")
